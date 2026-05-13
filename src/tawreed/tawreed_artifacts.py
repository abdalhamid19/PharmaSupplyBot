"""Artifact capture helpers for Tawreed Playwright failures."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Page

from ..core.artifact_run import artifact_filename, current_artifact_run


def dump_artifacts(page: Page, profile_key: str, label: str, details: str = "") -> None:
    """Persist screenshot, HTML, and URL artifacts for debugging failures."""
    try:
        artifacts_dir = _artifacts_dir(profile_key)
        screenshot_path = _artifact_path(profile_key, label, None, ".png")
        html_path = _artifact_path(profile_key, label, None, ".html")
        text_path = _artifact_path(profile_key, label, None, ".txt")
        _write_screenshot_artifact(page, screenshot_path)
        _write_html_artifact(page, html_path)
        _write_text_artifact(page, text_path, details)
        print(f"[{profile_key}] Saved artifacts to: {artifacts_dir}")
    except Exception:
        pass


def write_text_artifact(profile_key: str, label: str, content: str) -> Path:
    """Write a plain-text diagnostic artifact for the active profile."""
    artifact_path = _artifact_path(profile_key, label, None, ".txt")
    artifact_path.write_text(content, encoding="utf-8")
    return artifact_path


def append_text_artifact(
    profile_key: str,
    label: str,
    content: str,
    label_suffix: str | None = None,
) -> Path:
    """Append plain-text diagnostic content to a profile artifact."""
    artifact_path = _artifact_path(profile_key, label, label_suffix, ".txt")
    with artifact_path.open("a", encoding="utf-8") as artifact_file:
        artifact_file.write(content)
    return artifact_path


def append_csv_artifact(
    profile_key: str,
    label: str,
    rows: list[dict[str, object]],
    label_suffix: str | None = None,
) -> Path | None:
    """Append structured diagnostic rows to a CSV artifact for the active profile."""
    if not rows:
        return None
    artifact_path = _artifact_path(profile_key, label, label_suffix, ".csv")
    fieldnames = _csv_artifact_fieldnames(artifact_path, rows)
    _ensure_csv_schema(artifact_path, fieldnames)
    _append_csv_rows(artifact_path, fieldnames, rows)
    return artifact_path


def append_xlsx_artifact(
    profile_key: str,
    label: str,
    rows: list[dict[str, object]],
    label_suffix: str | None = None,
) -> Path | None:
    """Append structured rows to an XLSX artifact for the active profile."""
    if not rows:
        return None
    try:
        artifact_path = _artifact_path(profile_key, label, label_suffix, ".xlsx")
        workbook, worksheet = _open_or_create_workbook(artifact_path)
        fieldnames = _xlsx_artifact_fieldnames(worksheet, rows)
        _ensure_xlsx_header_row(worksheet, fieldnames)
        _append_xlsx_rows(worksheet, fieldnames, rows)
        workbook.save(artifact_path)
        return artifact_path
    except PermissionError:
        _log_xlsx_permission_error(profile_key, label)
        return None


def _log_xlsx_permission_error(profile_key: str, label: str) -> None:
    """Print a friendly message when an XLSX artifact is open elsewhere."""
    print(f"[{profile_key}] Could not update {label}.xlsx; close it and retry.")


def _artifact_path(
    profile_key: str, label: str, label_suffix: str | None, extension: str
) -> Path:
    """Return the fully-qualified artifact path for one label."""
    active = current_artifact_run()
    if active:
        return active.directory / artifact_filename(label, extension, label_suffix)
    effective_label = f"{label}.{label_suffix}" if label_suffix else label
    return _artifacts_dir(profile_key) / f"{effective_label}{extension}"


def _ensure_csv_schema(artifact_path: Path, fieldnames: list[str]) -> None:
    """Rewrite an existing CSV artifact when its schema has changed."""
    if not artifact_path.exists():
        return
    existing_fieldnames = _csv_header_fieldnames(artifact_path)
    if existing_fieldnames and existing_fieldnames != fieldnames:
        _rewrite_csv_artifact_with_fieldnames(artifact_path, fieldnames)


def _csv_artifact_fieldnames(
    artifact_path: Path, rows: list[dict[str, object]]
) -> list[str]:
    """Return existing CSV columns plus any new row keys in first-seen order."""
    fieldnames = _csv_header_fieldnames(artifact_path) if artifact_path.exists() else []
    return _merge_row_fieldnames(fieldnames, rows)


def _xlsx_artifact_fieldnames(worksheet, rows: list[dict[str, object]]) -> list[str]:
    """Return existing XLSX columns plus any new row keys in first-seen order."""
    return _merge_row_fieldnames(_merge_row_fieldnames([], rows), _xlsx_rows(worksheet))


def _merge_row_fieldnames(
    fieldnames: list[str], rows: list[dict[str, object]]
) -> list[str]:
    """Return fieldnames extended with keys that appear in the supplied rows."""
    merged = list(fieldnames)
    seen = set(merged)
    for row in rows:
        for key in row:
            if key not in seen:
                merged.append(key)
                seen.add(key)
    return merged


def _append_csv_rows(
    artifact_path: Path, fieldnames: list[str], rows: list[dict[str, object]]
) -> None:
    """Append rows to a CSV artifact with a header when needed."""
    write_header = not artifact_path.exists() or artifact_path.stat().st_size == 0
    with artifact_path.open("a", encoding="utf-8", newline="") as artifact_file:
        writer = csv.DictWriter(artifact_file, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        writer.writerows(rows)


def _append_xlsx_rows(
    worksheet, fieldnames: list[str], rows: list[dict[str, object]]
) -> None:
    """Append row values to an XLSX worksheet."""
    for row in rows:
        worksheet.append([row.get(field_name, "") for field_name in fieldnames])


def _artifacts_dir(profile_key: str) -> Path:
    """Return the artifacts directory for the active profile, creating it if needed."""
    active = current_artifact_run()
    artifacts_dir = active.directory if active else Path("artifacts") / profile_key
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    return artifacts_dir


def _open_or_create_workbook(artifact_path: Path):
    """Return an existing workbook or a new workbook for the requested artifact path."""
    if artifact_path.exists():
        workbook = load_workbook(artifact_path)
        return workbook, workbook.active
    workbook = Workbook()
    return workbook, workbook.active


def _csv_header_fieldnames(artifact_path: Path) -> list[str]:
    """Return the existing CSV header row fieldnames when the file is present."""
    with artifact_path.open("r", encoding="utf-8", newline="") as artifact_file:
        reader = csv.DictReader(artifact_file)
        return list(reader.fieldnames or [])


def _rewrite_csv_artifact_with_fieldnames(
    artifact_path: Path, fieldnames: list[str]
) -> None:
    """Rewrite an existing CSV artifact so all rows conform to the requested fieldnames."""
    with artifact_path.open("r", encoding="utf-8", newline="") as artifact_file:
        reader = csv.DictReader(artifact_file)
        existing_rows = list(reader)
    with artifact_path.open("w", encoding="utf-8", newline="") as artifact_file:
        writer = csv.DictWriter(artifact_file, fieldnames=fieldnames)
        writer.writeheader()
        for row in existing_rows:
            writer.writerow(
                {field_name: row.get(field_name, "") for field_name in fieldnames}
            )


def _ensure_xlsx_header_row(worksheet, fieldnames: list[str]) -> None:
    """Ensure the worksheet starts with exactly one header row."""
    if worksheet.max_row == 1 and worksheet.cell(row=1, column=1).value is None:
        worksheet.delete_rows(1, 1)
        worksheet.append(fieldnames)
        return
    existing_fieldnames = _xlsx_header_fieldnames(worksheet)
    if existing_fieldnames and existing_fieldnames != fieldnames:
        _rewrite_xlsx_worksheet_with_fieldnames(
            worksheet, existing_fieldnames, fieldnames
        )


def _xlsx_header_fieldnames(worksheet) -> list[str]:
    """Return the existing XLSX header fieldnames when a worksheet has a header row."""
    if worksheet.max_row == 0:
        return []
    values = [
        worksheet.cell(row=1, column=i).value
        for i in range(1, worksheet.max_column + 1)
    ]
    return [str(v) for v in values if v not in (None, "")]


def _xlsx_rows(worksheet) -> list[dict[str, object]]:
    """Return lightweight placeholder rows for existing XLSX columns."""
    fieldnames = _xlsx_header_fieldnames(worksheet)
    return [dict.fromkeys(fieldnames, "")]


def _rewrite_xlsx_worksheet_with_fieldnames(
    worksheet, existing_fieldnames: list[str], fieldnames: list[str]
) -> None:
    """Rewrite an existing worksheet so rows conform to the requested fieldnames."""
    existing_rows = []
    for row_values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            f: row_values[i] if i < len(row_values) else ""
            for i, f in enumerate(existing_fieldnames)
        }
        existing_rows.append(row)
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(fieldnames)
    for row in existing_rows:
        worksheet.append([row.get(f, "") for f in fieldnames])


def _write_screenshot_artifact(page: Page, screenshot_path: Path) -> None:
    """Write a full-page screenshot when artifact capture is requested."""
    try:
        page.screenshot(path=str(screenshot_path), full_page=True)
    except Exception:
        pass


def _write_html_artifact(page: Page, html_path: Path) -> None:
    """Write prettified HTML content for offline inspection."""
    try:
        html_content = page.content()
        pretty_html = html_content.replace("><", ">\n<")
        html_path.write_text(pretty_html, encoding="utf-8")
    except Exception:
        pass


def _write_text_artifact(page: Page, text_path: Path, details: str) -> None:
    """Write lightweight text diagnostics for the current page state."""
    try:
        content = f"url={page.url}\n"
        if details:
            content += details
        text_path.write_text(content, encoding="utf-8")
    except Exception:
        pass
