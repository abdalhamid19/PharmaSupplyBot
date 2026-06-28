"""XLSX artifact helpers for Tawreed."""

from __future__ import annotations

from openpyxl import Workbook, load_workbook


def _open_or_create_workbook(artifact_path):
    """Return an existing workbook or a new workbook for the requested artifact path."""
    if artifact_path.exists():
        workbook = load_workbook(artifact_path)
        return workbook, workbook.active
    workbook = Workbook()
    return workbook, workbook.active


def _xlsx_artifact_fieldnames(worksheet, rows: list[dict[str, object]]) -> list[str]:
    """Return existing XLSX columns plus any new row keys in first-seen order."""
    from .tawreed_artifacts_csv import _merge_row_fieldnames
    return _merge_row_fieldnames(_merge_row_fieldnames([], rows), _xlsx_rows(worksheet))


def _ensure_xlsx_header_row(worksheet, fieldnames: list[str]) -> None:
    """Ensure the worksheet starts with exactly one header row."""
    if worksheet.max_row == 1 and worksheet.cell(row=1, column=1).value is None:
        worksheet.delete_rows(1, 1)
        worksheet.append(fieldnames)
        return
    existing_fieldnames = _xlsx_header_fieldnames(worksheet)
    if existing_fieldnames and existing_fieldnames != fieldnames:
        _rewrite_xlsx_worksheet_with_fieldnames(
            worksheet, existing_fieldnames, fieldnames
        )


def _append_xlsx_rows(
    worksheet, fieldnames: list[str], rows: list[dict[str, object]]
) -> None:
    """Append row values to an XLSX worksheet."""
    for row in rows:
        worksheet.append([row.get(field_name, "") for field_name in fieldnames])


def _xlsx_header_fieldnames(worksheet) -> list[str]:
    """Return the existing XLSX header fieldnames when a worksheet has a header row."""
    if worksheet.max_row == 0:
        return []
    values = [
        worksheet.cell(row=1, column=i).value
        for i in range(1, worksheet.max_column + 1)
    ]
    return [str(v) for v in values if v not in (None, "")]


def _xlsx_rows(worksheet) -> list[dict[str, object]]:
    """Return lightweight placeholder rows for existing XLSX columns."""
    fieldnames = _xlsx_header_fieldnames(worksheet)
    return [dict.fromkeys(fieldnames, "")]


def _rewrite_xlsx_worksheet_with_fieldnames(
    worksheet, existing_fieldnames: list[str], fieldnames: list[str]
) -> None:
    """Rewrite an existing worksheet so rows conform to the requested fieldnames."""
    existing_rows = []
    for row_values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            f: row_values[i] if i < len(row_values) else ""
            for i, f in enumerate(existing_fieldnames)
        }
        existing_rows.append(row)
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(fieldnames)
    for row in existing_rows:
        worksheet.append([row.get(f, "") for f in fieldnames])
