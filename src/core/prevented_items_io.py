"""File I/O operations for prevented items."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .item_text import (
    display_code_text,
    normalized_cell_text,
)
from .prevented_items_constants import (
    DEFAULT_PREVENTED_ITEMS_PATH,
    PREVENTED_CODE_COLUMN,
    PREVENTED_NAME_COLUMN,
    PreventedItem,
)
from .prevented_items_helpers import (
    _prevented_column_indices,
    _prevented_item_from_row,
    normalized_prevented_key,
)


def load_prevented_items(
    path: Path = DEFAULT_PREVENTED_ITEMS_PATH,
) -> list[PreventedItem]:
    """Load prevented items from an XLSX file using openpyxl."""
    if not path.exists():
        return []
    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if sheet is None:
                return []
            rows = sheet.iter_rows(min_row=1, values_only=True)
            return _parse_prevented_rows(rows, path)
        finally:
            workbook.close()


def _parse_prevented_rows(rows, path) -> list[PreventedItem]:
    """Extract and validate prevented items from Excel rows."""
    code_idx, name_idx = _prevented_column_indices(next(rows), path)
    items: list[PreventedItem] = []
    seen_keys: set[tuple[str, str]] = set()
    for row in rows:
        item = _prevented_item_from_row(row, code_idx, name_idx)
        if item is None:
            continue
        key = normalized_prevented_key(item)
        if key not in seen_keys:
            seen_keys.add(key)
            items.append(item)
    return items


def save_prevented_items(
    prevented_items: list[PreventedItem],
    path: Path = DEFAULT_PREVENTED_ITEMS_PATH,
) -> Path:
    """Save the prevented list to disk as an XLSX file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Could not create prevented-items worksheet.")
    sheet.append([PREVENTED_CODE_COLUMN, PREVENTED_NAME_COLUMN])
    for item in prevented_items:
        sheet.append([item.code, item.name])
    workbook.save(str(path))
    return path


__all__ = [
    "load_prevented_items",
    "_parse_prevented_rows",
    "save_prevented_items",
    "normalized_prevented_key",
]
