"""Persistent prevented-item list helpers for Tawreed orders."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl

from .item_text import (
    display_code_text,
    normalize_code,
    normalize_name,
    normalized_cell_text,
    normalized_key,
)
from .utils.excel import Item

# Constants
PREVENTED_ITEMS_DIR = Path("data/input") / "prevented_items"
DEFAULT_PREVENTED_ITEMS_PATH = PREVENTED_ITEMS_DIR / "drugprevented.xlsx"
PREVENTED_CODE_COLUMN = "كود"
PREVENTED_NAME_COLUMN = "إسم الصنف"


@dataclass(frozen=True)
class PreventedItem:
    """One product that should not be ordered from Tawreed."""

    code: str
    name: str


# Helper functions
def _prevented_column_indices(header_row, path) -> tuple[int, int]:
    """Return prevented Excel code/name column indices."""
    header = [str(cell).strip() if cell else "" for cell in header_row]
    try:
        return header.index(PREVENTED_CODE_COLUMN), header.index(PREVENTED_NAME_COLUMN)
    except ValueError:
        raise KeyError(f"Missing columns in {path}. Found: {header}")


def _prevented_item_from_row(
    row: tuple[object, ...], code_idx: int, name_idx: int
) -> PreventedItem | None:
    """Return one prevented item from a raw Excel row when populated."""
    code = display_code_text(row[code_idx])
    name = normalized_cell_text(row[name_idx])
    if not code and not name:
        return None
    return PreventedItem(code=code, name=name)


def normalized_prevented_key(item: PreventedItem) -> tuple[str, str]:
    """Return the comparable identity for one prevented item."""
    return normalized_key(item.code, item.name)


def _normalized_path(path: Path) -> Path:
    """Return an absolute path for reliable same-file comparisons."""
    return path.expanduser().resolve(strict=False)


# File I/O operations
def load_prevented_items(
    path: Path = DEFAULT_PREVENTED_ITEMS_PATH,
) -> list[PreventedItem]:
    """Load prevented items from an XLSX file using openpyxl."""
    if not path.exists():
        return []
    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if sheet is None:
                return []
            rows = sheet.iter_rows(min_row=1, values_only=True)
            return _parse_prevented_rows(rows, path)
        finally:
            workbook.close()


def _parse_prevented_rows(rows, path) -> list[PreventedItem]:
    """Extract and validate prevented items from Excel rows."""
    code_idx, name_idx = _prevented_column_indices(next(rows), path)
    items: list[PreventedItem] = []
    seen_keys: set[tuple[str, str]] = set()
    for row in rows:
        item = _prevented_item_from_row(row, code_idx, name_idx)
        if item is None:
            continue
        key = normalized_prevented_key(item)
        if key not in seen_keys:
            seen_keys.add(key)
            items.append(item)
    return items


def save_prevented_items(
    prevented_items: list[PreventedItem],
    path: Path = DEFAULT_PREVENTED_ITEMS_PATH,
) -> Path:
    """Save the prevented list to disk as an XLSX file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Could not create prevented-items worksheet.")
    sheet.append([PREVENTED_CODE_COLUMN, PREVENTED_NAME_COLUMN])
    for item in prevented_items:
        sheet.append([item.code, item.name])
    workbook.save(str(path))
    return path


# Filtering operations
def add_prevented_item(
    prevented_items: list[PreventedItem],
    code: object,
    name: object,
) -> list[PreventedItem]:
    """Return the prevented list with one item added if it is not already present."""
    item = PreventedItem(code=display_code_text(code), name=normalized_cell_text(name))
    if not item.code and not item.name:
        return prevented_items
    existing_keys = {normalized_prevented_key(existing) for existing in prevented_items}
    if normalized_prevented_key(item) in existing_keys:
        return prevented_items
    return [*prevented_items, item]


def remove_prevented_item(
    prevented_items: list[PreventedItem],
    code: object,
    name: object,
) -> list[PreventedItem]:
    """Return the prevented list without the matching item."""
    remove_key = normalized_key(code, name)
    return [
        item for item in prevented_items if normalized_prevented_key(item) != remove_key
    ]


def filter_prevented_order_items(
    items: Iterable[Item],
    prevented_items: list[PreventedItem],
) -> Iterable[Item]:
    """Yield order items excluding any products in the prevented list."""
    blocked_codes = {normalize_code(item.code) for item in prevented_items if item.code}
    code_only_blocks = {
        normalize_code(item.code)
        for item in prevented_items
        if item.code and not item.name
    }
    blocked_names = {normalize_name(item.name) for item in prevented_items if item.name}

    for item in items:
        if not _is_prevented_order_item(
            item, blocked_codes, code_only_blocks, blocked_names
        ):
            yield item


def _is_prevented_order_item(
    item: Item,
    blocked_codes: set[str],
    code_only_blocks: set[str],
    blocked_names: set[str],
) -> bool:
    """Return whether one order item is blocked by the prevented list."""
    item_code, item_name = normalized_key(item.code, item.name)
    if item_name in blocked_names:
        return True
    if item_code and item_name:
        return item_code in code_only_blocks
    if item_code:
        return item_code in blocked_codes
    return False


def is_prevented_items_excel_path(
    excel_path: Path,
    prevented_items_path,
) -> bool:
    """Return whether the order source path points at the prevented-items list."""
    if prevented_items_path is None:
        prevented_items_path = DEFAULT_PREVENTED_ITEMS_PATH
    return _normalized_path(excel_path) == _normalized_path(prevented_items_path)


# Public exports
__all__ = [
    "PREVENTED_ITEMS_DIR",
    "DEFAULT_PREVENTED_ITEMS_PATH",
    "PREVENTED_CODE_COLUMN",
    "PREVENTED_NAME_COLUMN",
    "PreventedItem",
    "load_prevented_items",
    "save_prevented_items",
    "add_prevented_item",
    "remove_prevented_item",
    "filter_prevented_order_items",
    "is_prevented_items_excel_path",
    "normalized_prevented_key",
]
