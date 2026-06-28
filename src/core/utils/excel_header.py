"""Excel header detection and column resolution utilities."""

from pathlib import Path
from typing import Any, cast

import openpyxl

from ..config.config_models import ExcelConfig


def _detect_header_row(path: Path, config: ExcelConfig) -> int | None:
    """Find the likely header row in report-style exports using openpyxl."""
    aliases = _required_column_aliases(config)
    required = [
        aliases[config.code_col],
        aliases[config.name_col],
        aliases[config.qty_col],
    ]
    identity = [aliases[config.code_col], aliases[config.name_col]]
    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            return _detect_header_row_in_sheet(workbook.active, required, identity)
        finally:
            workbook.close()


def _detect_header_row_in_sheet(sheet, required, identity) -> int | None:
    """Find a complete or partial header row in one worksheet."""
    best_partial: int | None = None
    for row_index, row in enumerate(
        cast(Any, sheet).iter_rows(max_row=20, values_only=True)
    ):
        values = {_normalize_header(value) for value in row if value is not None}
        if all(values.intersection(column_aliases) for column_aliases in required):
            return row_index
        if best_partial is None and all(
            values.intersection(column_aliases) for column_aliases in identity
        ):
            best_partial = row_index
    return best_partial


def _resolve_col_indices(
    sheet, header_index, config, match_only: bool = False
) -> list[int]:
    """Map required Excel columns to their numeric sheet indices."""
    header_row = _sheet_header_row(sheet, header_index)
    col_map = _header_col_map(header_row)
    required = _required_columns(config, match_only)
    aliases = _required_column_aliases(config)
    indices = [_first_matching_index(col_map, aliases[column]) for column in required]
    missing = [col for col, idx in zip(required, indices) if idx is None]
    if missing:
        _raise_missing_columns(missing, required, header_row)
    return [idx for idx in indices if idx is not None]


def _sheet_header_row(sheet, header_index) -> list:
    """Return the detected header row values."""
    return list(
        sheet.iter_rows(
            min_row=header_index + 1, max_row=header_index + 1, values_only=True
        )
    )[0]


def _header_col_map(header_row) -> dict[str, int]:
    """Return normalized header labels mapped to sheet indices."""
    return {_normalize_header(cell): idx for idx, cell in enumerate(header_row) if cell}


def _required_columns(config: ExcelConfig, match_only: bool) -> list[str]:
    """Return the Excel columns required for the requested load mode."""
    if match_only:
        return [config.code_col, config.name_col]
    return [config.code_col, config.name_col, config.qty_col]


def _raise_missing_columns(missing: list[str], required: list[str], header_row) -> None:
    """Raise a detailed error for missing configured Excel columns."""
    found = [str(cell).strip() for cell in header_row if cell]
    raise ValueError(
        f"Missing required Excel columns: {missing}. "
        f"Expected: {required}. Found: {found}"
    )


def _required_column_aliases(config: ExcelConfig) -> dict[str, list[str]]:
    """Return accepted header labels for each configured Excel column."""
    return {
        config.code_col: _column_aliases(config.code_col, {"الكود", "كود الصنف"}),
        config.name_col: _column_aliases(config.name_col, {"اسم الصنف"}),
        config.qty_col: _column_aliases(
            config.qty_col,
            {"الكمية المطلوبة", "كمية الطلب", "الكمية", "كمية"},
        ),
    }


def _column_aliases(primary: str, alternatives: set[str]) -> list[str]:
    """Normalize one configured header plus known local report variants."""
    aliases = [_normalize_header(primary)]
    aliases.extend(
        alias
        for alias in sorted({_normalize_header(value) for value in alternatives})
        if alias not in aliases
    )
    return aliases


def _normalize_header(value: object) -> str:
    """Normalize Excel header cells for stable Arabic/English matching."""
    return " ".join(str(value).strip().split())


def _first_matching_index(col_map: dict[str, int], aliases: list[str]) -> int | None:
    """Return the first sheet index matching any accepted column alias."""
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None
