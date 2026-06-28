"""Excel reading utilities for Tawreed order items."""

from pathlib import Path
from typing import Any, Iterable, cast

import openpyxl

from ..config.config_models import ExcelConfig
from .excel_header import _detect_header_row, _resolve_col_indices
from .excel_conversion import _row_tuple_to_item


def load_items_from_excel(
    path: Path, config: ExcelConfig, limit: int = 0
) -> Iterable:
    """Yield order items from an Excel sheet using the configured column mapping."""
    yield from _limited_items(_read_item_rows(path, config), config, limit)


def load_match_only_items_from_excel(
    path: Path, config: ExcelConfig, limit: int = 0
) -> Iterable:
    """Yield match-only items from code/name catalog sheets without quantities."""
    yield from _limited_items(_read_match_only_rows(path, config), config, limit)


def _limited_items(
    rows: Iterable[tuple], config: ExcelConfig, limit: int = 0
) -> Iterable:
    """Yield converted rows until the optional processing limit is reached."""
    count = 0
    for row in rows:
        item = _row_tuple_to_item(row, config)
        if item is None:
            continue
        yield item
        count += 1
        if limit and count >= limit:
            break


def _read_item_rows(path: Path, config: ExcelConfig) -> Iterable[tuple]:
    """Yield only the configured Excel item columns as row tuples using openpyxl."""
    header_row_index = _detect_header_row(path, config)
    if header_row_index is None:
        header_row_index = 0

    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            sheet = cast(Any, workbook.active)
            indices = _resolve_col_indices(sheet, header_row_index, config)

            for row in sheet.iter_rows(min_row=header_row_index + 2, values_only=True):
                yield tuple(row[idx] for idx in indices)
        finally:
            workbook.close()


def _read_match_only_rows(path: Path, config: ExcelConfig) -> Iterable[tuple]:
    """Yield code, name, and a default quantity from two-column catalog sheets."""
    header_row_index = _detect_header_row(path, config) or 0
    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            sheet = cast(Any, workbook.active)
            indices = _resolve_col_indices(sheet, header_row_index, config, True)
            for row in sheet.iter_rows(min_row=header_row_index + 2, values_only=True):
                yield (row[indices[0]], row[indices[1]], 1)
        finally:
            workbook.close()
