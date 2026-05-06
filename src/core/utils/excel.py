"""Excel parsing utilities for Tawreed order items."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, cast

import openpyxl

from ..config.config_models import ExcelConfig


@dataclass(frozen=True)
class Item:
    """One requested product row loaded from the shortage Excel sheet."""

    code: str
    name: str
    qty: int


def load_items_from_excel(
    path: Path, config: ExcelConfig, limit: int = 0
) -> Iterable[Item]:
    """Yield order items from an Excel sheet using the configured column mapping."""
    yield from _limited_items(_read_item_rows(path, config), config, limit)


def load_match_only_items_from_excel(
    path: Path, config: ExcelConfig, limit: int = 0
) -> Iterable[Item]:
    """Yield match-only items from code/name catalog sheets without quantities."""
    yield from _limited_items(_read_match_only_rows(path, config), config, limit)


def _limited_items(
    rows: Iterable[tuple], config: ExcelConfig, limit: int = 0
) -> Iterable[Item]:
    """Yield converted rows until the optional processing limit is reached."""
    count = 0
    for row in rows:
        item = _row_tuple_to_item(row, config)
        if item is None:
            continue
        yield item
        count += 1
        if limit and count >= limit:
            break


def _read_item_rows(path: Path, config: ExcelConfig) -> Iterable[tuple]:
    """Yield only the configured Excel item columns as row tuples using openpyxl."""
    header_row_index = _detect_header_row(path, config)
    if header_row_index is None:
        header_row_index = 0

    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            sheet = cast(Any, workbook.active)
            indices = _resolve_col_indices(sheet, header_row_index, config)

            # Iterate rows starting after header
            for row in sheet.iter_rows(min_row=header_row_index + 2, values_only=True):
                yield tuple(row[idx] for idx in indices)
        finally:
            workbook.close()


def _read_match_only_rows(path: Path, config: ExcelConfig) -> Iterable[tuple]:
    """Yield code, name, and a default quantity from two-column catalog sheets."""
    header_row_index = _detect_header_row(path, config) or 0
    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            sheet = cast(Any, workbook.active)
            indices = _resolve_col_indices(sheet, header_row_index, config, True)
            for row in sheet.iter_rows(min_row=header_row_index + 2, values_only=True):
                yield (row[indices[0]], row[indices[1]], 1)
        finally:
            workbook.close()


def _resolve_col_indices(
    sheet, header_index, config, match_only: bool = False
) -> list[int]:
    """Map required Excel columns to their numeric sheet indices."""
    header_row = _sheet_header_row(sheet, header_index)
    col_map = _header_col_map(header_row)
    required = _required_columns(config, match_only)
    aliases = _required_column_aliases(config)
    indices = [_first_matching_index(col_map, aliases[column]) for column in required]
    missing = [col for col, idx in zip(required, indices) if idx is None]
    if missing:
        _raise_missing_columns(missing, required, header_row)
    return [idx for idx in indices if idx is not None]


def _sheet_header_row(sheet, header_index) -> list:
    """Return the detected header row values."""
    return list(
        sheet.iter_rows(
            min_row=header_index + 1, max_row=header_index + 1, values_only=True
        )
    )[0]


def _header_col_map(header_row) -> dict[str, int]:
    """Return normalized header labels mapped to sheet indices."""
    return {_normalize_header(cell): idx for idx, cell in enumerate(header_row) if cell}


def _required_columns(config: ExcelConfig, match_only: bool) -> list[str]:
    """Return the Excel columns required for the requested load mode."""
    if match_only:
        return [config.code_col, config.name_col]
    return [config.code_col, config.name_col, config.qty_col]


def _raise_missing_columns(missing: list[str], required: list[str], header_row) -> None:
    """Raise a detailed error for missing configured Excel columns."""
    found = [str(cell).strip() for cell in header_row if cell]
    raise ValueError(
        f"Missing required Excel columns: {missing}. "
        f"Expected: {required}. Found: {found}"
    )


def _detect_header_row(path: Path, config: ExcelConfig) -> int | None:
    """Find the likely header row in report-style exports using openpyxl."""
    aliases = _required_column_aliases(config)
    required = [
        aliases[config.code_col],
        aliases[config.name_col],
        aliases[config.qty_col],
    ]
    identity = [aliases[config.code_col], aliases[config.name_col]]
    with open(path, "rb") as f:
        workbook = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            return _detect_header_row_in_sheet(workbook.active, required, identity)
        finally:
            workbook.close()


def _detect_header_row_in_sheet(sheet, required, identity) -> int | None:
    """Find a complete or partial header row in one worksheet."""
    best_partial: int | None = None
    for row_index, row in enumerate(
        cast(Any, sheet).iter_rows(max_row=20, values_only=True)
    ):
        values = {_normalize_header(value) for value in row if value is not None}
        if all(values.intersection(column_aliases) for column_aliases in required):
            return row_index
        if best_partial is None and all(
            values.intersection(column_aliases) for column_aliases in identity
        ):
            best_partial = row_index
    return best_partial


def _required_column_aliases(config: ExcelConfig) -> dict[str, list[str]]:
    """Return accepted header labels for each configured Excel column."""
    return {
        config.code_col: _column_aliases(config.code_col, {"الكود", "كود الصنف"}),
        config.name_col: _column_aliases(config.name_col, {"اسم الصنف"}),
        config.qty_col: _column_aliases(
            config.qty_col,
            {"الكمية المطلوبة", "كمية الطلب", "الكمية", "كمية"},
        ),
    }


def _column_aliases(primary: str, alternatives: set[str]) -> list[str]:
    """Normalize one configured header plus known local report variants."""
    aliases = [_normalize_header(primary)]
    aliases.extend(
        alias
        for alias in sorted({_normalize_header(value) for value in alternatives})
        if alias not in aliases
    )
    return aliases


def _normalize_header(value: object) -> str:
    """Normalize Excel header cells for stable Arabic/English matching."""
    return " ".join(str(value).strip().split())


def _first_matching_index(col_map: dict[str, int], aliases: list[str]) -> int | None:
    """Return the first sheet index matching any accepted column alias."""
    for alias in aliases:
        if alias in col_map:
            return col_map[alias]
    return None


def _row_tuple_to_item(row: tuple, config: ExcelConfig) -> Item | None:
    """Convert one Excel row tuple to an order item when it passes filters."""
    code = str(row[0] or "").strip()
    name = str(row[1] or "").strip()
    quantity = _bounded_quantity(row[2], config)
    if not code and not name:
        return None
    if quantity < config.min_qty:
        return None
    return Item(code=code, name=name, qty=quantity)


def _bounded_quantity(value: Any, config: ExcelConfig) -> int:
    """Clamp the requested quantity to the configured Excel limits."""
    quantity = _to_int(value)
    return min(quantity, config.max_qty)


def _to_int(x: Any) -> int:
    """Convert a spreadsheet cell to an integer quantity with empty-safe fallbacks."""
    if x is None or (isinstance(x, float) and x != x):  # x != x is NaN check
        return 0
    try:
        return int(round(float(x)))
    except Exception:
        s = str(x).strip()
        if not s:
            return 0
        return int(float(s))
