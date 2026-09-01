"""Excel catalog loader for the Excel target source.

The Excel target is a pharmacy/vendor pricelist in the shape::

    صنف | سعر | الخصم

(or with an optional leading code column). Each row becomes one
:class:`TargetProduct` so the matching engine can treat the catalog the
same way it treats a Tawreed search response.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, cast

import openpyxl

from ..config.config_models import ExcelTargetConfig


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class TargetProduct:
    """One row of an Excel target catalog, normalised for matching."""

    code: str
    name: str
    price: float
    discount_percent: float
    raw: dict[str, Any] = field(default_factory=dict)

    def to_candidate_dict(self) -> dict[str, Any]:
        """Return the candidate dict shape consumed by the core matcher.

        The matcher reads ``productNameEn``, ``productName``,
        ``availableQuantity``, ``discountPercent`` and ``salePrice`` from
        each candidate. We populate those keys so the same scoring engine
        works on Excel catalog rows without modification.
        """
        return {
            "productNameEn": self.name,
            "productNameEnFallback": self.name,
            "productName": self.name,
            "availableQuantity": 1,
            "productsCount": 1,
            "discountPercent": float(self.discount_percent or 0.0),
            "salePrice": float(self.price or 0.0),
            "storeProductId": self.code or f"row:{abs(hash(self.name))}",
            "excelTarget": True,
            "excelTargetRaw": dict(self.raw),
        }


_HEADER_NORMALIZE_RE = re.compile(r"\s+")


def _normalize_header(value: object) -> str:
    """Normalize one Excel header cell for stable Arabic matching."""
    if value is None:
        return ""
    return _HEADER_NORMALIZE_RE.sub(" ", str(value).strip())


def load_target_catalog_from_excel(
    path: Path, config: ExcelTargetConfig
) -> list[TargetProduct]:
    """Load the Excel target catalog from ``path`` using ``config``.

    The loader auto-detects the header row by scanning the first
    ``HEADER_SCAN_LIMIT`` rows for the configured ``name_col`` /
    ``price_col`` / ``discount_col`` headers. When the configured
    ``header_row`` is set explicitly it takes precedence and is trusted
    as-is.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel target file not found: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _select_sheet(workbook, config)
        header_index = _resolve_header_index(sheet, config)
        column_indices = _resolve_column_indices(sheet, header_index, config)
        products: list[TargetProduct] = []
        for row in sheet.iter_rows(
            min_row=header_index + 2, values_only=True
        ):
            product = _row_to_product(row, column_indices, config)
            if product is not None:
                products.append(product)
        return products
    finally:
        workbook.close()


HEADER_SCAN_LIMIT = 10


def _select_sheet(workbook, config: ExcelTargetConfig):
    """Return the worksheet the catalog lives in."""
    if config.sheet:
        if config.sheet in workbook.sheetnames:
            return workbook[config.sheet]
        raise ValueError(
            f"Excel target sheet '{config.sheet}' not found. "
            f"Available sheets: {workbook.sheetnames}"
        )
    return workbook.active


def _resolve_header_index(sheet, config: ExcelTargetConfig) -> int:
    """Find the row index that contains the configured headers."""
    if config.header_row:
        return int(config.header_row)

    name_alias = _normalize_header(config.name_col)
    price_alias = _normalize_header(config.price_col)
    discount_alias = _normalize_header(config.discount_col)
    for row_index, row in enumerate(
        cast(Any, sheet).iter_rows(
            max_row=HEADER_SCAN_LIMIT, values_only=True
        )
    ):
        normalized = {_normalize_header(cell) for cell in row if cell is not None}
        if name_alias in normalized and (
            price_alias in normalized or discount_alias in normalized
        ):
            return row_index
    return 0


def _resolve_column_indices(
    sheet, header_index: int, config: ExcelTargetConfig
) -> dict[str, int]:
    """Map configured column names to numeric indices on the header row."""
    header_row = list(
        cast(Any, sheet).iter_rows(
            min_row=header_index + 1,
            max_row=header_index + 1,
            values_only=True,
        )
    )[0]
    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _normalize_header(cell)
        if key:
            col_map[key] = idx

    indices: dict[str, int] = {}
    for logical, configured in (
        ("name", config.name_col),
        ("price", config.price_col),
        ("discount", config.discount_col),
        ("code", config.code_col),
    ):
        if not configured:
            continue
        key = _normalize_header(configured)
        if key not in col_map:
            raise ValueError(
                f"Excel target column '{configured}' not found in header row. "
                f"Found headers: {sorted(col_map)}"
            )
        indices[logical] = col_map[key]
    return indices


def _row_to_product(
    row: tuple,
    indices: dict[str, int],
    config: ExcelTargetConfig,
) -> TargetProduct | None:
    """Convert one Excel row tuple into a TargetProduct."""
    name_cell = row[indices["name"]] if "name" in indices else None
    name = str(name_cell or "").strip()
    if not name:
        return None

    raw: dict[str, Any] = {"name": name}
    price = 0.0
    if "price" in indices:
        price = _to_float(row[indices["price"]])
        raw["price"] = price

    discount = 0.0
    if "discount" in indices:
        discount = _to_float(row[indices["discount"]])
        raw["discount"] = discount

    code = ""
    if "code" in indices:
        code = str(row[indices["code"]] or "").strip()
        raw["code"] = code
    if not code and config.requires_code:
        return None

    return TargetProduct(
        code=code,
        name=name,
        price=price,
        discount_percent=discount,
        raw=raw,
    )


def _to_float(value: Any) -> float:
    """Coerce one Excel cell to ``float``, treating empties as zero."""
    if value is None:
        return 0.0
    if isinstance(value, float) and value != value:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        text = str(value).strip()
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0


def iter_target_candidates(
    products: Iterable[TargetProduct],
) -> list[dict[str, Any]]:
    """Materialise an iterable of products into candidate dicts."""
    return [product.to_candidate_dict() for product in products]


__all__ = [
    "TargetProduct",
    "load_target_catalog_from_excel",
    "iter_target_candidates",
]