"""Artifact capture helpers for Tawreed Playwright failures."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Page


def dump_artifacts(page: Page, profile_key: str, label: str, details: str = "") -> None:
    """Persist screenshot, HTML, and URL artifacts for debugging failures."""
    try:
        artifacts_dir = _artifacts_dir(profile_key)
        screenshot_path = artifacts_dir / f"{label}.png"
        html_path = artifacts_dir / f"{label}.html"
        text_path = artifacts_dir / f"{label}.txt"
        _write_screenshot_artifact(page, screenshot_path)
        _write_html_artifact(page, html_path)
        _write_text_artifact(page, text_path, details)
        print(f"[{profile_key}] Saved artifacts to: {artifacts_dir}")
    except Exception:
        pass


def write_text_artifact(profile_key: str, label: str, content: str) -> Path:
    """Write a plain-text diagnostic artifact for the active profile."""
    artifacts_dir = _artifacts_dir(profile_key)
    artifact_path = artifacts_dir / f"{label}.txt"
    artifact_path.write_text(content, encoding="utf-8")
    return artifact_path


def append_text_artifact(profile_key: str, label: str, content: str) -> Path:
    """Append plain-text diagnostic content to a profile artifact."""
    artifacts_dir = _artifacts_dir(profile_key)
    artifact_path = artifacts_dir / f"{label}.txt"
    with artifact_path.open("a", encoding="utf-8") as artifact_file:
        artifact_file.write(content)
    return artifact_path


def append_csv_artifact(profile_key: str, label: str, rows: list[dict[str, object]]) -> Path | None:
    """Append structured diagnostic rows to a CSV artifact for the active profile."""
    if not rows:
        return None
    artifacts_dir = _artifacts_dir(profile_key)
    artifact_path = artifacts_dir / f"{label}.csv"
    fieldnames = list(rows[0].keys())
    if artifact_path.exists():
        existing_fieldnames = _csv_header_fieldnames(artifact_path)
        if existing_fieldnames and existing_fieldnames != fieldnames:
            _rewrite_csv_artifact_with_fieldnames(artifact_path, fieldnames)
    should_write_header = not artifact_path.exists() or artifact_path.stat().st_size == 0
    with artifact_path.open("a", encoding="utf-8", newline="") as artifact_file:
        writer = csv.DictWriter(artifact_file, fieldnames=fieldnames)
        if should_write_header:
            writer.writeheader()
        writer.writerows(rows)
    return artifact_path


def append_xlsx_artifact(profile_key: str, label: str, rows: list[dict[str, object]]) -> Path | None:
    """Append structured rows to an XLSX artifact for the active profile."""
    if not rows:
        return None
    try:
        artifacts_dir = _artifacts_dir(profile_key)
        artifact_path = artifacts_dir / f"{label}.xlsx"
        fieldnames = list(rows[0].keys())
        workbook, worksheet = _open_or_create_workbook(artifact_path)
        _ensure_xlsx_header_row(worksheet, fieldnames)
        for row in rows:
            worksheet.append([row.get(field_name, "") for field_name in fieldnames])
        workbook.save(artifact_path)
        return artifact_path
    except PermissionError:
        print(
            f"[{profile_key}] Could not update {label}.xlsx because it is open in another program."
        )
        return None


def _artifacts_dir(profile_key: str) -> Path:
    """Return the artifacts directory for the active profile, creating it if needed."""
    artifacts_dir = Path("artifacts") / profile_key
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    return artifacts_dir


def _open_or_create_workbook(artifact_path: Path):
    """Return an existing workbook or a new workbook for the requested artifact path."""
    if artifact_path.exists():
        workbook = load_workbook(artifact_path)
        return workbook, workbook.active
    workbook = Workbook()
    return workbook, workbook.active


def _csv_header_fieldnames(artifact_path: Path) -> list[str]:
    """Return the existing CSV header row fieldnames when the file is present."""
    with artifact_path.open("r", encoding="utf-8", newline="") as artifact_file:
        reader = csv.DictReader(artifact_file)
        return list(reader.fieldnames or [])


def _rewrite_csv_artifact_with_fieldnames(artifact_path: Path, fieldnames: list[str]) -> None:
    """Rewrite an existing CSV artifact so all rows conform to the requested fieldnames."""
    with artifact_path.open("r", encoding="utf-8", newline="") as artifact_file:
        reader = csv.DictReader(artifact_file)
        existing_rows = list(reader)
    with artifact_path.open("w", encoding="utf-8", newline="") as artifact_file:
        writer = csv.DictWriter(artifact_file, fieldnames=fieldnames)
        writer.writeheader()
        for row in existing_rows:
            writer.writerow({field_name: row.get(field_name, "") for field_name in fieldnames})


def _ensure_xlsx_header_row(worksheet, fieldnames: list[str]) -> None:
    """Ensure the worksheet starts with exactly one header row."""
    if worksheet.max_row == 1 and worksheet.cell(row=1, column=1).value is None:
        worksheet.delete_rows(1, 1)
        worksheet.append(fieldnames)


def _write_screenshot_artifact(page: Page, screenshot_path: Path) -> None:
    """Write a full-page screenshot when artifact capture is requested."""
    try:
        page.screenshot(path=str(screenshot_path), full_page=True)
    except Exception:
        pass


def _write_html_artifact(page: Page, html_path: Path) -> None:
    """Write prettified HTML content for offline inspection."""
    try:
        html_content = page.content()
        pretty_html = html_content.replace("><", ">\n<")
        html_path.write_text(pretty_html, encoding="utf-8")
    except Exception:
        pass


def _write_text_artifact(page: Page, text_path: Path, details: str) -> None:
    """Write lightweight text diagnostics for the current page state."""
    try:
        content = f"url={page.url}\n"
        if details:
            content += details
        text_path.write_text(content, encoding="utf-8")
    except Exception:
        pass
