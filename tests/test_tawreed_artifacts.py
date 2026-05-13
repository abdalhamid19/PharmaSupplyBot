import csv
import os
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from src.tawreed.tawreed_artifacts import append_csv_artifact, append_xlsx_artifact


class TawreedArtifactsTests(unittest.TestCase):
    def test_append_csv_artifact_keeps_union_schema_for_mixed_rows(self) -> None:
        with TemporaryDirectory() as temp_dir:
            original_cwd = Path.cwd()
            try:
                os.chdir(temp_dir)
                append_csv_artifact(
                    "wardany",
                    "order_ai_trace",
                    [
                        {"phase": "ai_final", "item_code": "1", "ai_status": "ok"},
                        {
                            "phase": "api_attempt_verify",
                            "item_code": "1",
                            "provider": "groq",
                            "status": 429,
                        },
                    ],
                )
                append_csv_artifact(
                    "wardany",
                    "order_ai_trace",
                    [{"phase": "api_attempt_review", "model": "m1"}],
                )
                path = Path("artifacts") / "wardany" / "order_ai_trace.csv"
                with path.open("r", encoding="utf-8", newline="") as f:
                    reader = csv.DictReader(f)
                    fieldnames = reader.fieldnames
                    rows = list(reader)
            finally:
                os.chdir(original_cwd)

        self.assertEqual(
            fieldnames,
            ["phase", "item_code", "ai_status", "provider", "status", "model"],
        )
        self.assertEqual(rows[0]["ai_status"], "ok")
        self.assertEqual(rows[1]["provider"], "groq")
        self.assertEqual(rows[2]["model"], "m1")

    def test_append_xlsx_artifact_rewrites_changed_headers(self) -> None:
        with TemporaryDirectory() as temp_dir:
            original_cwd = Path.cwd()
            try:
                os.chdir(temp_dir)
                artifact_dir = Path("artifacts") / "wardany"
                artifact_dir.mkdir(parents=True)
                artifact_path = artifact_dir / "order_result_summary.xlsx"
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.append(["item_code", "matched_query", "searched_queries_count"])
                worksheet.append(["123", "Panadol", 2])
                workbook.save(artifact_path)

                append_xlsx_artifact(
                    "wardany",
                    "order_result_summary",
                    [
                        {
                            "item_code": "456",
                            "matched_query": "Panadol Advance",
                            "selected_discount_percent": "35%",
                            "selected_store_name": "Abu Amira",
                            "searched_queries_count": 3,
                        }
                    ],
                )

                loaded = load_workbook(artifact_path)
                values = list(loaded.active.iter_rows(values_only=True))
            finally:
                os.chdir(original_cwd)

        self.assertEqual(
            values[0],
            (
                "item_code",
                "matched_query",
                "selected_discount_percent",
                "selected_store_name",
                "searched_queries_count",
            ),
        )
        self.assertEqual(values[1], ("123", "Panadol", None, None, 2))
        self.assertEqual(values[2], ("456", "Panadol Advance", "35%", "Abu Amira", 3))


if __name__ == "__main__":
    unittest.main()
