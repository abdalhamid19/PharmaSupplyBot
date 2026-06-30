"""Tests for Tawreed product catalog export helpers."""

from __future__ import annotations

import csv
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

from openpyxl import load_workbook

from src.tawreed.products.tawreed_product_export import (
    ARABIC_EXPORT_SEARCH_TERMS,
    ENGLISH_EXPORT_SEARCH_TERMS,
    EXPORT_FIELDNAMES,
    EXPORT_SEARCH_TERMS,
    ProductExportRow,
    ProductSearchRequest,
    collect_unique_product_candidates,
    iter_all_product_candidates,
    iter_product_search_candidates,
    product_export_rows,
    write_product_export_files,
)


class TawreedProductExportTests(unittest.TestCase):
    """Validate Tawreed product export normalization, pagination, and files."""

    def test_product_export_rows_normalize_and_deduplicate_candidates(self) -> None:
        candidates = [
            {
                "productName": "  بانادول  ",
                "productNameEn": " Panadol ",
                "storeProductId": 123,
            },
            {
                "productName": "بانادول",
                "productNameEn": "Panadol",
                "storeProductId": 123,
            },
            {},
        ]

        rows = list(product_export_rows(candidates))

        self.assertEqual(rows, [ProductExportRow("بانادول", "Panadol", "123")])

    def test_product_export_rows_extract_sale_price(self) -> None:
        candidates = [
            {
                "productName": "بانادول",
                "productNameEn": "Panadol",
                "storeProductId": 123,
                "salePrice": 42.5,
            },
        ]

        rows = list(product_export_rows(candidates))

        self.assertEqual(rows[0].sale_price, "42.5")

    def test_product_export_rows_prefers_retail_price_for_sale_price(self) -> None:
        candidates = [
            {
                "productName": "بانادول ادفانس 500 مجم 48 اقراص",
                "productNameEn": "PANADOL ADVANCE 500 MG 48 F.C.TABS.",
                "storeProductId": 1653038,
                "salePrice": 64.39,
                "retailPrice": 92.0,
                "discountPercent": 30.01,
            },
        ]

        rows = list(product_export_rows(candidates))

        self.assertEqual(rows[0].sale_price, "92.0")

    def test_product_export_rows_keep_product_id_without_store_id(self) -> None:
        candidates = [
            {
                "productId": 34,
                "productName": "ابيمول 300 مجم 5 لبوس",
                "productNameEn": "ABIMOL 300 MG 5 RECTAL SUPP.",
                "storeProductId": None,
                "salePrice": 15.0,
            },
        ]

        rows = list(product_export_rows(candidates))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].product_id, "34")
        self.assertEqual(rows[0].store_product_id, "")

    def test_write_product_export_files_creates_all_requested_formats(self) -> None:
        rows = [ProductExportRow("بانادول", "Panadol", "123", sale_price="42.5")]

        with TemporaryDirectory() as temp_dir:
            paths = write_product_export_files(rows, Path(temp_dir), "catalog")
            csv_rows = _read_csv_rows(paths["csv"])
            txt_lines = paths["txt"].read_text(encoding="utf-8").splitlines()
            xlsx_rows = _read_xlsx_rows(paths["xlsx"])

        self.assertEqual(set(paths.keys()), {"csv", "xlsx", "txt"})
        self.assertEqual(csv_rows[0], list(EXPORT_FIELDNAMES))
        self.assertEqual(csv_rows[1][EXPORT_FIELDNAMES.index("sale_price")], "42.5")
        self.assertEqual(txt_lines[1].split("\t"), rows[0].values())
        self.assertEqual(xlsx_rows[0], EXPORT_FIELDNAMES)
        self.assertEqual(xlsx_rows[1][EXPORT_FIELDNAMES.index("sale_price")], "42.5")

    def test_iter_all_product_candidates_pages_until_total_pages(self) -> None:
        page = _FakePage([_payload(2, "1"), _payload(2, "2")])

        candidates = list(iter_all_product_candidates(page, page_size=10))

        self.assertEqual([row["storeProductId"] for row in candidates], ["1", "2"])
        self.assertEqual(page.request.urls[0].split("page=")[1].split("&")[0], "0")
        self.assertEqual(page.request.urls[1].split("page=")[1].split("&")[0], "1")

    def test_iter_all_product_candidates_honors_limit(self) -> None:
        page = _FakePage([_payload(3, "1", "2")])

        candidates = list(iter_all_product_candidates(page, page_size=10, limit=1))

        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["storeProductId"], "1")

    def test_export_search_terms_are_general_english_then_arabic(self) -> None:
        self.assertEqual(EXPORT_SEARCH_TERMS[0], "")
        self.assertEqual(EXPORT_SEARCH_TERMS[1:27], ENGLISH_EXPORT_SEARCH_TERMS)
        self.assertEqual(EXPORT_SEARCH_TERMS[27:], ARABIC_EXPORT_SEARCH_TERMS)
        self.assertIn("أ", ARABIC_EXPORT_SEARCH_TERMS)
        self.assertIn("ة", ARABIC_EXPORT_SEARCH_TERMS)
        self.assertIn("ى", ARABIC_EXPORT_SEARCH_TERMS)

    def test_iter_product_search_candidates_pages_each_request_in_order(self) -> None:
        page = _FakePage([_payload(2, "general-1"), _payload(2, "general-2"),
                          _payload(1, "a-1")])
        searches = iter([
            ProductSearchRequest("", {"x-test": "1"}, {"term": ""}),
            ProductSearchRequest("A", {"x-test": "1"}, {"term": "A"}),
        ])

        candidates = list(iter_product_search_candidates(page, searches, page_size=10))

        self.assertEqual([row["storeProductId"] for row in candidates],
                         ["general-1", "general-2", "a-1"])
        self.assertEqual([body["term"] for body in page.request.bodies],
                         ["", "", "A"])
        self.assertEqual([url.split("page=")[1].split("&")[0]
                          for url in page.request.urls], ["0", "1", "0"])

    def test_collect_unique_product_candidates_deduplicates_search_results(self) -> None:
        candidates = [
            _candidate("Panadol", "بنادول", "1"),
            _candidate("Panadol", "بنادول", "1"),
            _candidate("Aspirin", "اسبرين", "2"),
        ]

        collection = collect_unique_product_candidates(candidates)

        self.assertEqual([row["storeProductId"] for row in collection.candidates],
                         ["1", "2"])
        self.assertEqual(collection.scanned_count, 3)
        self.assertEqual(collection.duplicates_removed, 1)

    def test_collect_unique_product_candidates_limits_final_unique_rows(self) -> None:
        candidates = [
            _candidate("Panadol", "بنادول", "1"),
            _candidate("Panadol", "بنادول", "1"),
            _candidate("Aspirin", "اسبرين", "2"),
        ]

        collection = collect_unique_product_candidates(candidates, limit=1)

        self.assertEqual([row["storeProductId"] for row in collection.candidates],
                         ["1"])
        self.assertEqual(collection.scanned_count, 1)
        self.assertEqual(collection.duplicates_removed, 0)


def _read_csv_rows(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as input_file:
        return list(csv.reader(input_file))


def _read_xlsx_rows(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True)
    try:
        return list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def _payload(total_pages: int, *store_product_ids: str) -> dict[str, Any]:
    return {
        "data": {
            "totalPages": total_pages,
            "content": [
                {
                    "productName": "عربي",
                    "productNameEn": "English",
                    "storeProductId": value,
                }
                for value in store_product_ids
            ],
        }
    }


def _candidate(name_en: str, name_ar: str, store_id: str) -> dict[str, Any]:
    return {
        "productNameEn": name_en,
        "productName": name_ar,
        "storeProductId": store_id,
    }


class _FakePage:
    def __init__(self, payloads: list[dict[str, Any]]) -> None:
        self.request = _FakeRequest(payloads)

    def evaluate(self, expression: str) -> str:
        return "https://seller.tawreed.io"


class _FakeRequest:
    def __init__(self, payloads: list[dict[str, Any]]) -> None:
        self.payloads = payloads
        self.urls: list[str] = []
        self.bodies: list[dict[str, Any]] = []

    def post(
        self, url: str, data: Any, headers: dict[str, str], timeout: int
    ) -> "_FakeResponse":
        self.urls.append(url)
        self.bodies.append(data)
        return _FakeResponse(self.payloads.pop(0))


class _FakeResponse:
    ok = True
    status = 200

    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload

    def json(self) -> dict[str, Any]:
        return self.payload


if __name__ == "__main__":
    unittest.main()
